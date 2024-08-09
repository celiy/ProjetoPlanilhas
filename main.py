import xlsxwriter
from openpyxl import load_workbook
import pandas as pd
import os

colunas = []
def armazenar_colunas(worksheet):
    for cell in worksheet[1]:
        colunas.append(cell.value)

def arquivo(info_colunas, dadosp):

    file_path = 'planilha.xlsx'

    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        worksheet = workbook.active
        armazenar_colunas(worksheet)
    else:
        workbook = xlsxwriter.Workbook(file_path)
        worksheet = workbook.add_worksheet()

        rowschars = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        for c, p in enumerate(info_colunas):
            rowc = rowschars[c]+"1"
            worksheet.write(rowc, p)

        workbook.close()
        workbook = load_workbook(file_path)
        worksheet = workbook.active
        armazenar_colunas(worksheet)

    if dadosp is not None:
        #Adicionar as informações na planilha
        wb = load_workbook(file_path)
        ws = wb.active
        ws.append(dadosp)
        wb.save(file_path)

    workbook.close()

def mostrarplanilha():

    df = pd.read_excel("planilha.xlsx")

    print(df)

def carregarplanilha():
    fpath = 'planilha.xlsx'
    if os.path.exists(fpath):
        book = load_workbook(fpath)
        sheet = book.active
        armazenar_colunas(sheet)

carregarplanilha()

print("O que deseja fazer?")
print("1 - Ler planilha \n2 - Inserir dados \n3 - Modificar dados "
      "\n4 - Criar planilha")

opcao = int(input())
if opcao == 1:
    mostrarplanilha()

elif opcao == 2:
    print("Deseja inserir novas colunas/linhas (1) ou dados a elas (2)? ")
    opcao = int(input())
    if opcao == 1:
        print()
    if opcao == 2:
        dados = []
        for titulo in colunas:
            print(f"Digite o dado para a coluna '{titulo}': ")
            dado = str(input())
            dados.append(dado)
        arquivo(None, dados)

elif opcao == 3:
    print("Deseja modiciar uma linha (1) ou coluna (2)?")
    opcao = int(input())
    if opcao == 1:
        print("TTT")
    elif opcao == 2:
        print("TTT")

elif opcao == 4:
    path = 'planilha.xlsx'
    if os.path.exists(path):
        print("Erro: A planilha já existe.")
    else:
        print("Quantas colunas deseja criar? (max: 25)")
        cols = int(input())
        print("Quais devem ser o titulo de cada coluna? ")
        titulo_colunas = []
        for n in range(cols):
            print(f"Titulo da coluna {n+1}: ")
            titulo_colunas.append(str(input()))
        print(f"Dados inseridos: {titulo_colunas}")
        arquivo(titulo_colunas,None)