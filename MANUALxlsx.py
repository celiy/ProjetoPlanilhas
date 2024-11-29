try:
    import xlsxwriter
    from openpyxl import load_workbook
    from collections import Counter
    import pandas as pd
    import os
    import tkinter as tk
    from tkinter import filedialog
except Exception as error:
    print(error)
    print("Houve um erro na hora de importar a dependências. Por favor, os instale com o comando a seguir no seu CMD/Terminal: "
    "\n pip install -r requirements.txt \n\nSe isto não funcionar, tente executar o arquivo normalmente com duplo click invés do terminal.")
    input()
else:

    colunas = []

    def mostrarplanilha(filepath):
        df = pd.read_excel(filepath)
        print(df)
        input("Pressione qualquer tecla para continuar")

    def armazenar_colunas(worksheet):
        for cell in worksheet[1]:
            colunas.append(cell.value)

    def carregarplanilha(filepath):
        fpath = filepath
        if os.path.exists(fpath):
            book = load_workbook(fpath)
            sheet = book.active
            armazenar_colunas(sheet)
            return True
        else:
            return False

    def editarelemento(filepath):
        workbook = load_workbook(filepath)
        sheet = workbook.active

        print("Modificando planilha...")
        print("Insira a coluna: (A, B, C, D, ...)")
        coluna = str(input())
        coluna = coluna.upper()
        print("Insira a linha: ")
        linha = int(input())
        linha = linha+2
        print("Insira o valor novo: ")
        valor = str(input())
        celula = sheet[coluna + str(linha)]
        if celula.value is None:
            print("Linha e coluna fora dos valores existentes")
        else:
            sheet[f'{coluna}{linha}'] = valor
            workbook.save(filepath)

        print("Dados atualizados: ")
        mostrarplanilha(filepath)

    def editararquivo(filepath):
        dados = []
        for titulo in colunas:
            print(f"Digite o dado para a coluna '{titulo}': ")
            dado = str(input())
            dados.append(dado)
        wb = load_workbook(filepath)
        ws = wb.active
        ws.append(dados)
        wb.save(filepath)

        print("Dados atualizados: ")
        mostrarplanilha(filepath)

    def arquivo(filepath, info_colunas):

        if os.path.exists(filepath):
            workbook = load_workbook(filepath)
            worksheet = workbook.active
            armazenar_colunas(worksheet)
        else:
            workbook = xlsxwriter.Workbook(filepath)
            worksheet = workbook.add_worksheet()

            rowschars = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
            for c, p in enumerate(info_colunas):
                rowc = rowschars[c]+"1"
                worksheet.write(rowc, p)

            workbook.close()
            workbook = load_workbook(filepath)
            worksheet = workbook.active
            armazenar_colunas(worksheet)

        print("Dados atualizados: ")
        mostrarplanilha(filepath)

        workbook.close()

    def criararquivo():
        print("Insira o nome do arquivo da planilha: ")
        global filepath
        filepath = str(input()) + ".xlsx"
        if os.path.exists(filepath):
            print("Erro: A planilha já existe.")
        else:
            print("Quantas colunas deseja criar? (max: 25)")
            cols = int(input())
            print("Quais devem ser o titulo de cada coluna? ")
            titulo_colunas = []
            for n in range(cols):
                print(f"Titulo da coluna {n + 1}: ")
                titulo_colunas.append(str(input()))
            print(f"Dados inseridos: {titulo_colunas}")
            arquivo(titulo_colunas)

            print("Dados atualizados: ")
            mostrarplanilha(filepath)

    def escanearplanilha(filepath):
        print("Deseja fazer qual tipo de escaneamento?\n"
            "1 - Expressão matematica \n"
            "2 - Valores repetidos \n"
            "3 - Maior/Menor valor \n")
        opc = int(input())
        mostrarplanilha(filepath)
        rowcol = None
        while True:
            print("Deseja selecionar: linha (1) | coluna (2)")
            while True:
                try:
                    oprc = int(input())
                except ValueError:
                    print("Valor inválido, tente novamente: ")
                else:
                    break
            if oprc == 1:
                print("Insira a linha: (0, 1, 2, ...)")
                rowcol = int(input())
                rowcol = rowcol+2
                break
            elif oprc == 2:
                print("Insira a coluna: (A, B, C, ...)")
                rowcol = str(input())
                rowcol = rowcol.upper()
                break
            else:
                print("Inválido, tente novamente: ")

        workbook = load_workbook(filepath)
        sheet = workbook.active
        dados = []
        for cell in sheet[rowcol]:
            dados.append(cell.value)

        res = 0
        if opc == 1:
            while True:
                print("Qual tipo de expressão? \n"
                    "1 - Soma \n"
                    "2 - Multiplicação \n"
                    "3 - Média (divisão) \n"
                    "4 - Sair")
                op = int(input())
                if op == 1:
                    for n in dados:
                        if n.isdigit():
                            res += float(n)
                    print(f"A soma é: {res}")
                    res = 0
                elif op == 2:
                    res = 1
                    for n in dados:
                        if n.isdigit():
                            res = res * float(n)
                    print(f"A multiplicação é: {res}")
                    res = 0
                elif op == 3:
                    res = 0
                    for n in dados:
                        res += float(n)
                    res = res/len(dados)
                    print(f"A média é: {res}")
                    res = 0
                elif op == 4:
                    break

        elif opc == 2:
            contagem = Counter(dados)
            valor_mais_comum = contagem.most_common(1)[0][0]
            print(f"O valor mais comum é {valor_mais_comum}")
            input("Pressione qualquer tecla para continuar")

        elif opc == 3:
            print(f"O maior valor é {max(dados)}")
            print(f"O menor valor é {min(dados)}")
            input("Pressione qualquer tecla para continuar")

    #======================#
    #[ Começo do programa ]#
    #======================#
    def MANUALxlsx_main():
        def selecionar_arquivo():
            root = tk.Tk()
            root.withdraw()

            caminho_arquivo = filedialog.askopenfilename(title="Selecione a planilha")
                
            return caminho_arquivo

        filepath = selecionar_arquivo()

        if filepath:
            print(f"Arquivo selecionado: {filepath}")
        else:
            print("Nenhum arquivo selecionado.")

        while True:
            if carregarplanilha(filepath):
                break
            else:
                print("Nome inválido. Deseja criar uma planilha nova? "
                    "Sim (1) | Não (2)")
                opcao = int(input())
                if opcao == 1:
                    criararquivo()
                    exit = False
                    break
                else:
                    exit = True
                    break
    
        while not exit:
            print("O que deseja fazer?")
            print("1 - Ler planilha \n"
                "2 - Inserir dados \n"
                "3 - Modificar dados \n"
                "4 - Escanear dados \n"
                "5 - Criar planilha \n"
                "6 - Sair")

            while True:
                try:
                    opcao = int(input())
                except ValueError:
                    print("Valor inválido, tente novamente: ")
                else:
                    break

            if opcao == 1:
                mostrarplanilha(filepath)

            elif opcao == 2:
                editararquivo(filepath)

            elif opcao == 3:
                mostrarplanilha(filepath)
                editarelemento(filepath)

            elif opcao == 4:
                escanearplanilha(filepath)

            elif opcao == 5:
                criararquivo()

            elif opcao == 6:
                break